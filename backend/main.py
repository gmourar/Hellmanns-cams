import asyncio, base64, io, json, os, uuid, logging
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, HTTPException, Depends, Header, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, and_

from db import init_db, get_db, Session as SessionModel, SessionLocal
import command_queue as q_module
import storage
from indexing_service import indexar_sessao
from rekognition_service import RekognitionService, COLLECTION_ID

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

AGENT_TOKEN = os.environ.get("AGENT_TOKEN", "changeme")

_default_origins = "http://localhost:3000,http://127.0.0.1:3000"
ALLOWED_ORIGINS = [
    o.strip()
    for o in os.environ.get("ALLOWED_ORIGINS", _default_origins).split(",")
    if o.strip()
]


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    logger.info("Database initialized")
    await _recover_pending_sessions()
    yield


async def _recover_pending_sessions():
    """
    Recupera sessões com estado inconsistente após crash/restart:
    - status='recording' → reenfileira o RECORD (agente nunca recebeu)
    - indexing_status='indexing' ou status='ready'+pending → relança indexação
    """
    _asyncio = asyncio
    async with SessionLocal() as db:
        result = await db.execute(
            select(SessionModel).where(
                or_(
                    SessionModel.status == "recording",
                    SessionModel.indexing_status == "indexing",
                    and_(
                        SessionModel.status == "ready",
                        SessionModel.indexing_status == "pending",
                    ),
                )
            )
        )
        sessions = result.scalars().all()

    for session in sessions:
        if session.status == "recording":
            if session.agent_acked_at is not None:
                # Agent already received the command — trust it to complete or timeout
                logger.warning(
                    "Sessão %s em 'recording' mas agente já deu ack em %s — aguardando conclusão",
                    session.session_id, session.agent_acked_at,
                )
                continue
            logger.warning("Recuperando sessão %s em 'recording' — reenfileirando RECORD", session.session_id)
            await q_module.push({"type": "RECORD", "session_id": session.session_id})
        elif session.status == "ready" and session.indexing_status in ("indexing", "pending"):
            logger.warning("Relançando indexação interrompida da sessão %s", session.session_id)
            _asyncio.create_task(indexar_sessao(session.session_id))

    if sessions:
        logger.info("Recuperação: %d sessão(ões) reprocessada(s)", len(sessions))
    else:
        logger.info("Nenhuma sessão pendente para recuperar")


app = FastAPI(title="Hellmann's Cam Backend", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Auth ─────────────────────────────────────────────────────────────────────

def verify_agent(authorization: str = Header(...)):
    if authorization != f"Bearer {AGENT_TOKEN}":
        raise HTTPException(status_code=401, detail="Invalid agent token")


# ── Schemas ───────────────────────────────────────────────────────────────────

class CreateSessionRequest(BaseModel):
    operator_name: str
    participants: list[str] = []


class CompleteSessionRequest(BaseModel):
    status: str                  # "ok" | "error"
    detail: str | None = None
    cabine_ids: list[int] = []   # cabines that uploaded successfully


class GalleryVideo(BaseModel):
    cabine_id: int
    video_url: str
    qr_url: str


class BuscarVideoRequest(BaseModel):
    imagem_base64: str  # JPEG/PNG em base64


class BuscarVideoResponse(BaseModel):
    session_id: str
    cabine_id: int
    video_url: str
    similarity: float


# ── Operator endpoints ────────────────────────────────────────────────────────

@app.post("/operator/sessions", status_code=201)
async def create_session(
    body: CreateSessionRequest,
    db: AsyncSession = Depends(get_db),
):
    session_id = str(uuid.uuid4())[:8]
    session = SessionModel(
        session_id=session_id,
        operator_name=body.operator_name,
        participants=json.dumps(body.participants),
        status="recording",
        created_at=datetime.now(timezone.utc),
    )
    db.add(session)
    await db.commit()
    await q_module.push({"type": "RECORD", "session_id": session_id})
    logger.info("Session %s created by %s", session_id, body.operator_name)
    return {"session_id": session_id, "status": "recording"}


# ── Agent endpoints ───────────────────────────────────────────────────────────

@app.post("/agent/sessions/{session_id}/ack")
async def agent_ack(
    session_id: str,
    db: AsyncSession = Depends(get_db),
    _=Depends(verify_agent),
):
    """Agent calls this immediately after receiving a RECORD command to prevent double-dispatch on restart."""
    result = await db.execute(
        select(SessionModel).where(SessionModel.session_id == session_id)
    )
    session = result.scalar_one_or_none()
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")
    session.agent_acked_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


@app.get("/agent/sessions/current")
async def agent_current_session(db: AsyncSession = Depends(get_db), _=Depends(verify_agent)):
    """Retorna a sessão mais recente em status 'recording', sem consumir a fila."""
    result = await db.execute(
        select(SessionModel)
        .where(SessionModel.status == "recording")
        .order_by(SessionModel.created_at.desc())
        .limit(1)
    )
    session = result.scalar_one_or_none()
    if session is None:
        return Response(status_code=204)
    return {"session_id": session.session_id}


@app.get("/agent/poll")
async def agent_poll(_=Depends(verify_agent)):
    command = await q_module.wait(timeout=30.0)
    if command is None:
        from fastapi.responses import Response
        return Response(status_code=204)
    return command


@app.post("/agent/sessions/{session_id}/complete")
async def agent_complete(
    session_id: str,
    body: CompleteSessionRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    _=Depends(verify_agent),
):
    result = await db.execute(
        select(SessionModel).where(SessionModel.session_id == session_id)
    )
    session = result.scalar_one_or_none()
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")

    if body.status == "ok":
        session.status = "ready"
        if body.cabine_ids:
            session.cabine_ids = json.dumps(sorted(body.cabine_ids))
        await db.commit()
        background_tasks.add_task(indexar_sessao, session_id)
    else:
        session.status = "error"
        logger.error("Session %s error: %s", session_id, body.detail)
        await db.commit()

    logger.info("Session %s → %s", session_id, session.status)
    return {"ok": True}


# ── Gallery endpoint ──────────────────────────────────────────────────────────

@app.get("/gallery/{session_id}")
async def gallery(session_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(SessionModel).where(SessionModel.session_id == session_id)
    )
    session = result.scalar_one_or_none()
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")

    participants = session.participants_list
    cabine_ids = session.cabine_ids_list
    if not cabine_ids:
        cabine_ids = await storage.available_cabine_ids_async(session_id)

    _asyncio = asyncio
    urls = await _asyncio.gather(*[
        storage.public_url_async(session_id, c) for c in cabine_ids
    ])
    videos = [
        GalleryVideo(
            cabine_id=c,
            video_url=url,
            qr_url=f"{storage.BASE_URL}/gallery/{session_id}/cabines/{c}/qr.svg",
        )
        for c, url in zip(cabine_ids, urls)
    ]

    return {
        "session_id": session_id,
        "participants": participants,
        "videos": videos,
        "video_urls": [video.video_url for video in videos],
        "status": session.status,
        "indexing_status": session.indexing_status or "pending",
        "created_at": (session.created_at.isoformat() + "Z") if session.created_at else None,
    }


# ── QR endpoint ───────────────────────────────────────────────────────────────

@app.get("/gallery/{session_id}/cabines/{cabine_id}/qr.svg")
async def cabine_qr(session_id: str, cabine_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(SessionModel).where(SessionModel.session_id == session_id)
    )
    session = result.scalar_one_or_none()
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found")
    if not await storage.video_exists_async(session_id, cabine_id):
        raise HTTPException(status_code=404, detail="Video not found")

    import qrcode
    import qrcode.image.svg

    # Stable redirect URL — never expires, generates fresh presigned on each scan
    stable_url = f"{storage.BASE_URL}/videos/{session_id}/cabine_{cabine_id}.mp4"
    qr = qrcode.QRCode(border=2)
    qr.add_data(stable_url)
    qr.make(fit=True)
    image = qr.make_image(image_factory=qrcode.image.svg.SvgPathImage)
    buffer = io.BytesIO()
    image.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="image/svg+xml",
        headers={"Cache-Control": "public, max-age=300"},
    )


# ── QR genérico do evento ─────────────────────────────────────────────────────

@app.get("/meu-video/qr.svg")
async def meu_video_qr():
    import qrcode
    import qrcode.image.svg

    frontend_url = os.environ.get("FRONTEND_BASE_URL", "http://localhost:3000")
    face_scan_url = f"{frontend_url}/meu-video"

    qr = qrcode.QRCode(border=2)
    qr.add_data(face_scan_url)
    qr.make(fit=True)
    image = qr.make_image(image_factory=qrcode.image.svg.SvgPathImage)
    buffer = io.BytesIO()
    image.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="image/svg+xml",
        headers={"Cache-Control": "public, max-age=3600"},
    )


# ── Buscar vídeo por reconhecimento facial ────────────────────────────────────

@app.post("/buscar-video")
async def buscar_video(body: BuscarVideoRequest):
    try:
        image_bytes = base64.b64decode(body.imagem_base64)
    except Exception:
        raise HTTPException(status_code=400, detail="imagem_base64 inválida")

    if not image_bytes:
        raise HTTPException(status_code=400, detail="Imagem vazia. Tente tirar a foto novamente.")

    rekognition = RekognitionService()
    try:
        matches = await rekognition.buscar_rosto_async(image_bytes, COLLECTION_ID)
    except Exception:
        raise HTTPException(status_code=400, detail="Não foi possível detectar um rosto. Centralize seu rosto e tente novamente.")

    if not matches:
        raise HTTPException(status_code=404, detail="Rosto não encontrado. Tente novamente.")

    # Deduplica por external_image_id (vários frames da mesma cabine podem dar match)
    seen: set[str] = set()
    candidates: list[tuple[str, int, float]] = []
    for match in sorted(matches, key=lambda m: m["similarity"], reverse=True):
        external_id = match["external_image_id"]
        if external_id in seen:
            continue
        seen.add(external_id)
        try:
            parts = external_id.split("_c")
            s_id = parts[0]
            c_id = int(parts[1])
        except (IndexError, ValueError):
            continue
        candidates.append((s_id, c_id, match["similarity"]))

    # Check S3 existence and generate URLs in parallel
    _asyncio = asyncio
    exists_flags, urls = await _asyncio.gather(
        _asyncio.gather(*[storage.video_exists_async(s, c) for s, c, _ in candidates]),
        _asyncio.gather(*[storage.public_url_async(s, c) for s, c, _ in candidates]),
    )

    results = []
    for (s_id, c_id, similarity), ok, url in zip(candidates, exists_flags, urls):
        if not ok:
            continue
        results.append({
            "session_id": s_id,
            "cabine_id": c_id,
            "video_url": url,
            "similarity": similarity,
        })

    if not results:
        raise HTTPException(status_code=404, detail="Vídeo não encontrado no storage")

    return results


# ── List all sessions ─────────────────────────────────────────────────────────

@app.get("/sessions")
async def list_sessions(
    db: AsyncSession = Depends(get_db),
    page: int = 1,
    limit: int = 20,
):
    offset = (page - 1) * limit
    result = await db.execute(
        select(SessionModel)
        .where(SessionModel.status == "ready")
        .order_by(SessionModel.created_at.desc())
        .offset(offset)
        .limit(limit + 1)  # +1 to check if there are more
    )
    sessions_raw = result.scalars().all()
    has_more = len(sessions_raw) > limit
    sessions = sessions_raw[:limit]

    async def _session_videos(session: SessionModel) -> list[dict]:
        cabine_ids = session.cabine_ids_list or []
        if cabine_ids:
            # Verify S3 existence even for stored IDs (handles deleted files)
            exists_flags = await asyncio.gather(*[
                storage.video_exists_async(session.session_id, c) for c in cabine_ids
            ])
            cabine_ids = [c for c, ok in zip(cabine_ids, exists_flags) if ok]
        if not cabine_ids:
            cabine_ids = await storage.available_cabine_ids_async(session.session_id)
        if not cabine_ids:
            return []
        urls = await asyncio.gather(*[
            storage.public_url_async(session.session_id, c) for c in cabine_ids
        ])
        return [
            {
                "cabine_id": c,
                "video_url": url,
                "qr_url": f"{storage.BASE_URL}/gallery/{session.session_id}/cabines/{c}/qr.svg",
            }
            for c, url in zip(cabine_ids, urls)
        ]

    _asyncio = asyncio
    video_lists = await _asyncio.gather(*[_session_videos(s) for s in sessions])

    response = []
    for session, videos in zip(sessions, video_lists):
        response.append({
            "session_id": session.session_id,
            "operator_name": session.operator_name,
            "participants": session.participants_list,
            "created_at": (session.created_at.isoformat() + "Z") if session.created_at else None,
            "indexing_status": session.indexing_status or "pending",
            "videos": videos,
        })

    # Filter out sessions with deleted/missing videos
    response = [s for s in response if s["videos"]]

    return {"sessions": response, "has_more": has_more}


# ── Video file serving ────────────────────────────────────────────────────────

@app.get("/videos/{session_id}/{filename}")
async def serve_video(session_id: str, filename: str):
    if storage.STORAGE_BACKEND == "s3":
        cabine_id = None
        if filename.startswith("cabine_") and filename.endswith(".mp4"):
            try:
                cabine_id = int(filename.removeprefix("cabine_").removesuffix(".mp4"))
            except ValueError:
                pass
        if cabine_id and await storage.video_exists_async(session_id, cabine_id):
            from fastapi.responses import RedirectResponse
            return RedirectResponse(await storage.public_url_async(session_id, cabine_id))
        raise HTTPException(status_code=404, detail="Video on S3 — use gallery URLs")
    path = storage.video_dir(session_id) / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail="Video not found")
    return FileResponse(str(path), media_type="video/mp4")


# ── Admin report (all-time, per day) ─────────────────────────────────────────

@app.get("/admin/report")
async def admin_report(db: AsyncSession = Depends(get_db)):
    """Retorna sessões e vídeos agrupados por dia — a partir de 03/06/2026 (SP)."""
    SP_OFFSET = timedelta(hours=-3)
    EVENT_START_UTC = datetime(2026, 6, 3, 3, 0, 0)  # 03/06 00:00 SP = 03:00 UTC

    result = await db.execute(
        select(SessionModel)
        .where(SessionModel.status == "ready")
        .where(SessionModel.created_at >= EVENT_START_UTC)
        .order_by(SessionModel.created_at)
    )
    sessions = result.scalars().all()

    from collections import defaultdict
    by_day: dict = defaultdict(lambda: {"sessoes": 0, "videos": 0})
    for s in sessions:
        if s.created_at is None:
            continue
        sp_dt = s.created_at + SP_OFFSET
        day = sp_dt.strftime("%Y-%m-%d")
        by_day[day]["sessoes"] += 1
        by_day[day]["videos"] += len(s.cabine_ids_list)

    rows = [
        {"data": d, "sessoes": v["sessoes"], "videos": v["videos"]}
        for d, v in sorted(by_day.items())
    ]
    return {
        "dias": rows,
        "total_sessoes": sum(r["sessoes"] for r in rows),
        "total_videos": sum(r["videos"] for r in rows),
        "gerado_em": datetime.utcnow().isoformat() + "Z",
    }


@app.get("/admin/report/excel")
async def admin_report_excel(db: AsyncSession = Depends(get_db)):
    """Gera e retorna o relatório Excel para download — a partir de 03/06/2026 (SP)."""
    import io as _io
    from collections import defaultdict
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    SP_OFFSET = timedelta(hours=-3)
    EVENT_START_UTC = datetime(2026, 6, 3, 3, 0, 0)
    DAYS_PT = {0:"Segunda",1:"Terça",2:"Quarta",3:"Quinta",4:"Sexta",5:"Sábado",6:"Domingo"}

    result = await db.execute(
        select(SessionModel)
        .where(SessionModel.status == "ready")
        .where(SessionModel.created_at >= EVENT_START_UTC)
        .order_by(SessionModel.created_at)
    )
    sessions = result.scalars().all()

    by_day: dict = defaultdict(lambda: {"sessoes": 0, "videos": 0})
    for s in sessions:
        if s.created_at is None:
            continue
        sp_dt = s.created_at + SP_OFFSET
        day = sp_dt.strftime("%Y-%m-%d")
        by_day[day]["sessoes"] += 1
        by_day[day]["videos"] += len(s.cabine_ids_list)

    YELLOW, BLACK, WHITE, LIGHT, GRAY = "FFDD00", "1A1A1A", "FFFFFF", "FFF9CC", "F5F5F5"

    def bthin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def bmed():
        s = Side(style="medium", color=BLACK)
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sessões por Dia"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Hellmann's — Ativação Fotográfica"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BLACK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"Relatório gerado em {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC"
    c.font = Font(name="Calibri", size=10, color="666666")
    c.fill = PatternFill("solid", fgColor=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    for col, h in enumerate(["Data", "Dia da Semana", "Sessões Realizadas", "Total de Vídeos"], 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, size=11, color=BLACK)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bthin()
    ws.row_dimensions[4].height = 22

    sorted_days = sorted(by_day.items())
    for i, (day, v) in enumerate(sorted_days):
        row = 5 + i
        dt = datetime.strptime(day, "%Y-%m-%d")
        fill = PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
        for col, val in enumerate([
            dt.strftime("%d/%m/%Y"),
            DAYS_PT[dt.weekday()],
            v["sessoes"],
            v["videos"],
        ], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", size=11)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bthin()
        ws.row_dimensions[row].height = 20

    tr = 5 + len(sorted_days)
    ws.row_dimensions[tr].height = 22
    for col, val in [(1, "TOTAL"), (2, f"{len(sorted_days)} dias")]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLACK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bthin()
    for col, val in [(3, sum(v["sessoes"] for v in by_day.values())),
                     (4, sum(v["videos"] for v in by_day.values()))]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=12, color=BLACK)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bmed()

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"participantes-{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Admin dashboard stats ─────────────────────────────────────────────────────

@app.get("/admin/stats")
async def admin_stats(db: AsyncSession = Depends(get_db)):
    SP_OFFSET = timedelta(hours=-3)  # UTC-3 Sao Paulo
    # Event start: 2026-06-02 00:00 SP = 2026-06-02 03:00 UTC
    EVENT_START_UTC = datetime(2026, 6, 2, 3, 0, 0)

    result = await db.execute(
        select(SessionModel).where(SessionModel.created_at >= EVENT_START_UTC)
    )
    sessions = result.scalars().all()

    now_utc = datetime.utcnow()
    now_sp = now_utc + SP_OFFSET
    today_sp_midnight_utc = datetime(now_sp.year, now_sp.month, now_sp.day) - SP_OFFSET

    def is_today(s: SessionModel) -> bool:
        return bool(s.created_at and s.created_at >= today_sp_midnight_utc)

    ready = [s for s in sessions if s.status == "ready"]
    ready_today = [s for s in ready if is_today(s)]
    recording_now = [s for s in sessions if s.status == "recording"]

    total_videos = sum(len(s.cabine_ids_list) for s in ready)
    videos_today = sum(len(s.cabine_ids_list) for s in ready_today)

    # Sessions per hour — last 24h, labeled in SP time
    sessions_per_hour = []
    for i in range(24):
        hour_start = (now_utc - timedelta(hours=23 - i)).replace(minute=0, second=0, microsecond=0)
        hour_end = hour_start + timedelta(hours=1)
        label = (hour_start + SP_OFFSET).strftime("%H:%M")
        count = sum(
            1 for s in ready
            if s.created_at and hour_start <= s.created_at < hour_end
        )
        sessions_per_hour.append({"hour": label, "count": count})

    # Videos per day — last 7 days, labeled in SP time
    videos_per_day = []
    for i in range(7):
        day_sp = now_sp.date() - timedelta(days=6 - i)
        day_start_utc = datetime(day_sp.year, day_sp.month, day_sp.day) - SP_OFFSET
        day_end_utc = day_start_utc + timedelta(days=1)
        day_sessions = [
            s for s in ready
            if s.created_at and day_start_utc <= s.created_at < day_end_utc
        ]
        videos_per_day.append({
            "day": day_sp.strftime("%d/%m"),
            "videos": sum(len(s.cabine_ids_list) for s in day_sessions),
            "sessions": len(day_sessions),
        })

    # Last 10 sessions (any status), most recent first
    recent = sorted(sessions, key=lambda x: x.created_at or datetime.min, reverse=True)[:10]
    recent_list = [
        {
            "session_id": s.session_id,
            "operator_name": s.operator_name,
            "created_at": (s.created_at.isoformat() + "Z") if s.created_at else None,
            "status": s.status,
            "video_count": len(s.cabine_ids_list),
            "participant_count": len(s.participants_list),
        }
        for s in recent
    ]

    return {
        "sessions_today": len(ready_today),
        "videos_today": videos_today,
        "total_sessions": len(ready),
        "total_videos": total_videos,
        "recording_now": len(recording_now),
        "sessions_per_hour": sessions_per_hour,
        "videos_per_day": videos_per_day,
        "recent_sessions": recent_list,
        "updated_at": now_utc.isoformat() + "Z",
    }
